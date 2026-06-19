"""Branded Excel renderer (openpyxl). Items become a live, formula-driven sheet."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from crm.documents.domain.enums import DOCUMENT_TYPE_LABELS


def _argb(hex_color: str, fallback: str = "FF7C6CFF") -> str:
    h = (hex_color or "").lstrip("#")
    return ("FF" + h.upper()) if len(h) == 6 else fallback


def _num_format(currency: str) -> str:
    return f'"{currency} "#,##0.00' if currency else "#,##0.00"


def render_xlsx(brand: dict, doc_type: str, payload: dict) -> bytes:
    label = DOCUMENT_TYPE_LABELS.get(doc_type, "Documento")
    currency = payload["currency"] or brand.get("currency", "")
    primary = _argb(brand.get("primary_color", "#7C6CFF"))
    accent = _argb(brand.get("accent_color", "#22C55E"), "FF22C55E")

    wb = Workbook()
    ws = wb.active
    ws.title = label[:31]

    white_bold = Font(bold=True, color="FFFFFFFF", size=12)
    head_font = Font(bold=True, color="FFFFFFFF", size=10)
    bold = Font(bold=True, size=10)
    muted = Font(color="FF6B6B7B", size=9)
    primary_fill = PatternFill("solid", fgColor=primary)
    accent_fill = PatternFill("solid", fgColor=accent)
    thin = Side(style="thin", color="FFE4E4EE")
    border = Border(bottom=thin)
    money_fmt = _num_format(currency)

    widths = {"A": 46, "B": 12, "C": 16, "D": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Header band.
    ws.merge_cells("A1:D1")
    ws["A1"] = brand.get("business_name") or "Documento"
    ws["A1"].font = white_bold
    ws["A1"].fill = primary_fill
    ws["A1"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    ws["A2"] = label + (f" · {payload['document_number']}" if payload["document_number"] else "")
    ws["A2"].font = bold
    ws["A3"] = payload["title"]
    if payload["subtitle"]:
        ws["A4"] = payload["subtitle"]
        ws["A4"].font = muted

    row = 6
    c = payload["client"]
    if c["name"]:
        ws[f"A{row}"] = f"Para: {c['name']}"
        ws[f"A{row}"].font = bold
        row += 1
    for line in [c["contact"], c["email"], c["phone"], c["tax_id"]]:
        if line:
            ws[f"A{row}"] = line
            ws[f"A{row}"].font = muted
            row += 1
    if payload["valid_until"]:
        ws[f"A{row}"] = f"Válido hasta: {payload['valid_until']}"
        ws[f"A{row}"].font = muted
        row += 1
    row += 1

    if payload["has_items"]:
        headers = ["Detalle", "Cantidad", "Precio unitario", "Importe"]
        for i, title in enumerate(headers):
            cell = ws.cell(row=row, column=i + 1, value=title)
            cell.font = head_font
            cell.fill = primary_fill
            cell.alignment = Alignment(horizontal="right" if i else "left", vertical="center")
        ws.row_dimensions[row].height = 20
        first_item = row + 1

        r = first_item
        for it in payload["items"]:
            ws.cell(row=r, column=1, value=it["description"] or "—").border = border
            ws.cell(row=r, column=2, value=float(it["quantity"])).border = border
            up = ws.cell(row=r, column=3, value=float(it["unit_price"]))
            up.number_format = money_fmt
            up.border = border
            tot = ws.cell(row=r, column=4, value=f"=B{r}*C{r}")
            tot.number_format = money_fmt
            tot.border = border
            r += 1
        last_item = r - 1

        # Totals with live formulas.
        sub_row = r + 1
        ws.cell(row=sub_row, column=3, value="Subtotal").alignment = Alignment(horizontal="right")
        ws.cell(row=sub_row, column=3).font = bold
        sub = ws.cell(row=sub_row, column=4, value=f"=SUM(D{first_item}:D{last_item})")
        sub.number_format = money_fmt
        sub.font = bold

        total_row = sub_row + 1
        if payload["tax_rate"] and payload["tax"]:
            tax_row = sub_row + 1
            total_row = sub_row + 2
            rate = float(payload["tax_rate"])
            ws.cell(row=tax_row, column=3, value=f"IVA ({rate:g}%)").alignment = Alignment(
                horizontal="right"
            )
            tax = ws.cell(row=tax_row, column=4, value=f"=D{sub_row}*{rate}/100")
            tax.number_format = money_fmt
            grand = f"=D{sub_row}+D{tax_row}"
        else:
            grand = f"=D{sub_row}"

        tcell_label = ws.cell(row=total_row, column=3, value="Total")
        tcell_label.font = Font(bold=True, color="FFFFFFFF")
        tcell_label.fill = accent_fill
        tcell_label.alignment = Alignment(horizontal="right")
        tcell = ws.cell(row=total_row, column=4, value=grand)
        tcell.number_format = money_fmt
        tcell.font = Font(bold=True, color="FFFFFFFF")
        tcell.fill = accent_fill
        row = total_row + 2

    # Narrative sections (works for reports and for itemized docs with context).
    for section in payload["sections"]:
        if section["heading"]:
            ws.cell(row=row, column=1, value=section["heading"]).font = bold
            row += 1
        if section["body"]:
            cell = ws.cell(row=row, column=1, value=section["body"])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 2

    terms = payload["terms"] or brand.get("default_terms", "")
    if terms:
        ws.cell(row=row, column=1, value="Términos y condiciones").font = bold
        row += 1
        ws.cell(row=row, column=1, value=terms).font = muted

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
