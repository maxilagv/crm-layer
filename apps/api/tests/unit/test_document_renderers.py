"""Document renderers produce valid PDF/XLSX/PPTX bytes from a payload."""

from decimal import Decimal
from io import BytesIO

import pytest

from crm.documents.domain.payload import normalize_payload
from crm.documents.renderers import render_pdf, render_pptx, render_xlsx

BRAND = {
    "business_name": "Mi Estudio",
    "legal_name": "Mi Estudio SRL",
    "tax_id": "30-11111111-9",
    "email": "hola@miestudio.com",
    "phone": "+54 11 5555-5555",
    "website": "miestudio.com",
    "address": "Av. Siempreviva 742",
    "primary_color": "#7C6CFF",
    "accent_color": "#22C55E",
    "text_color": "#16161D",
    "currency": "ARS",
    "default_terms": "Condiciones por defecto.",
    "footer_note": "Gracias por confiar.",
    "logo_bytes": None,
}

RAW_PAYLOAD = {
    "title": "Propuesta de desarrollo",
    "subtitle": "App de gestión de turnos",
    "client": {"name": "ACME SRL", "contact": "Juan", "email": "juan@acme.com", "tax_id": "30-9"},
    "intro": "Gracias por la oportunidad de trabajar juntos.",
    "sections": [{"heading": "Alcance", "body": "Backend + app web + soporte."}],
    "items": [
        {
            "description": "Desarrollo backend",
            "quantity": 1,
            "unit_price": "150000",
            "unit": "proyecto",
        },
        {"description": "Soporte mensual", "quantity": 3, "unit_price": "20000"},
    ],
    "currency": "ARS",
    "tax_rate": "21",
    "terms": "50% para arrancar, 50% a la entrega.",
    "valid_until": "2026-07-01",
    "meta": {"document_number": "P-0001"},
}


def test_normalize_payload_computes_totals():
    p = normalize_payload(RAW_PAYLOAD)
    assert p["subtotal"] == Decimal("210000.00")
    assert p["tax"] == Decimal("44100.00")
    assert p["total"] == Decimal("254100.00")
    assert p["has_items"] is True
    assert len(p["items"]) == 2
    assert p["items"][1]["line_total"] == Decimal("60000.00")


def test_render_pdf_is_valid():
    data = render_pdf(BRAND, "proposal", normalize_payload(RAW_PAYLOAD))
    assert isinstance(data, bytes) and data[:4] == b"%PDF"
    assert len(data) > 1500


def test_render_xlsx_is_valid():
    data = render_xlsx(BRAND, "quote", normalize_payload(RAW_PAYLOAD))
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data))
    ws = wb.active
    assert ws["A1"].value == "Mi Estudio"
    # The total cell must hold a live formula, not a literal.
    formulas = [
        c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]
    assert any("SUM" in f for f in formulas)


def test_render_pptx_is_valid():
    data = render_pptx(BRAND, "deck", normalize_payload(RAW_PAYLOAD))
    from pptx import Presentation

    prs = Presentation(BytesIO(data))
    # cover + 1 section + items slide
    assert len(prs.slides) >= 2


def test_renderers_handle_empty_report():
    payload = normalize_payload(
        {"title": "Informe", "sections": [{"heading": "Resumen", "body": "Todo ok."}]}
    )
    assert render_pdf(BRAND, "report", payload)[:4] == b"%PDF"
    assert render_xlsx(BRAND, "report", payload)
    assert render_pptx(BRAND, "report", payload)


@pytest.mark.parametrize("renderer", [render_pdf, render_xlsx, render_pptx])
def test_renderers_handle_minimal_payload(renderer):
    # No items, no sections, no client: must not crash.
    assert renderer(BRAND, "proposal", normalize_payload({"title": "X"}))
